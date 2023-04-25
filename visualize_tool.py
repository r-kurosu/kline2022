import time, datetime
import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import make_instance_tool
import MASTER

# color setting
yellow_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFFF00')
red_fill = excel.styles.PatternFill(patternType='solid', fgColor='FF0000')
blue_fill = excel.styles.PatternFill(patternType='solid', fgColor='0000FF')
green_fill = excel.styles.PatternFill(patternType='solid', fgColor='00FF00')
pink_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFC0CB')
skyblue_fill = excel.styles.PatternFill(patternType='solid', fgColor='87CEEB')
rime_fill = excel.styles.PatternFill(patternType='solid', fgColor='FAEBD7')
gray_fill = excel.styles.PatternFill(patternType='solid', fgColor='808080')
purple_fill = excel.styles.PatternFill(patternType='solid', fgColor='800080')
orange_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFA500')
yellow_green_fill = excel.styles.PatternFill(patternType='solid', fgColor='9ACD32')
pink_gold_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFD700')
wine_red_fill = excel.styles.PatternFill(patternType='solid', fgColor='800000')
marine_blue_fill = excel.styles.PatternFill(patternType='solid', fgColor='000080')
passion_pink_fill = excel.styles.PatternFill(patternType='solid', fgColor='FF1493')
gold_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFD700')
blue_white_fill = excel.styles.PatternFill(patternType='solid', fgColor='ADD8E6')

color_list = [yellow_green_fill, red_fill, blue_white_fill, pink_fill, blue_fill, gold_fill, green_fill, skyblue_fill, rime_fill, gray_fill, purple_fill, orange_fill, yellow_fill, pink_gold_fill, wine_red_fill, passion_pink_fill]

block_border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))

ramp_border = Border(top=Side(style='mediumDashed', color='FF0000'), 
                bottom=Side(style='mediumDashed', color='FF0000'), 
                left=Side(style='mediumDashed', color='FF0000'),
                right=Side(style='mediumDashed', color='FF0000'))

center_alignment = Alignment(horizontal="centerContinuous", vertical="center" , wrap_text=True)
red_font = excel.styles.fonts.Font(color='FF0000')
brue_font = excel.styles.fonts.Font(color='0000FF')


def paint_cell(a, b, sol, ws, m):
    enter_block, exit_block = make_instance_tool.get_ramp_block(a, b)
    
    for i in range(a):
        for j in range(b):
            ws.cell(row=2*i+1, column=2*j+1).value = i*b+j
            ws.cell(row=2*i+1, column=2*j+1).border = block_border
            ws.cell(row=2*i+1, column=2*j+1).alignment = center_alignment
            
            if i*b+j == enter_block or i*b+j == exit_block:
                continue
            else:
                ws.cell(row=2*i+1, column=2*j+1).fill = color_list[sol[i*b+j]]
    
    # dummy car
    try:
        ws.cell(row=2*(exit_block//b)+1, column=2*(exit_block%b)+1).fill = color_list[m]
    except:
        print("車種が多すぎます. カラーを追加してください")
        return ws
    
    return ws


def point_enter_cell(a, b, in_out_flag, ws):
    enter_block, exit_block = make_instance_tool.get_ramp_block(a, b)
    ws.cell(row=2*(enter_block//b)+1, column=2*(enter_block%b)+1).border = ramp_border
    ws.cell(row=2*(exit_block//b)+1, column=2*(exit_block%b)+1).border = ramp_border
    
    return ws


def print_edge_in(a, b, sola, ws):
    for edge in sola:
        if edge[0] + 1 == edge[1]:
            # print(f"edge →: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = (cell_contents or "") + "→"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge ←: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = (cell_contents or "") + "←"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    return


def print_edge_out(a, b, solb, ws):
    for edge in solb:
        if edge[0] + 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = (cell_contents or "") + "→"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = (cell_contents or "") + "←"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    
    return


def add_annotation(ws, a, b, m, LP_list, DP_list, Amount_list, M_h_bar):
    Hold_list = [x for x in M_h_bar if x not in [0, 1, 2, 3]]
        
    ws.cell(row=1, column=2*b+1).value = f"car"
    ws.cell(row=1, column=2*b+1).alignment = center_alignment
    ws.cell(row=1, column=2*b+2).value = f"LP → DP"
    ws.cell(row=1, column=2*b+2).alignment = center_alignment
    ws.cell(row=1, column=2*b+3).value = f"car amount (RT)"
    ws.cell(row=1, column=2*b+3).alignment = center_alignment
    ws.cell(row=1, column=2*b+4).value = f"hold"
    ws.cell(row=1, column=2*b+4).alignment = center_alignment
    
    for i in range(m):
        ws.cell(row=i+2, column=2*b+1).fill = color_list[i]
        ws.cell(row=i+2, column=2*b+1).value = f"car {i}"
        ws.cell(row=i+2, column=2*b+1).alignment = center_alignment
        ws.cell(row=i+2, column=2*b+2).value = f"{LP_list[i]} → {DP_list[i]}"
        ws.cell(row=i+2, column=2*b+2).alignment = center_alignment
        ws.cell(row=i+2, column=2*b+3).value = f"{Amount_list[i]}"
        ws.cell(row=i+2, column=2*b+3).alignment = center_alignment
        # ws.cell(row=i+2, column=2*b+4).value = f"{Hold_list[i]}"
        # ws.cell(row=i+2, column=2*b+4).alignment = center_alignment
    
    # dummy car
    if m >= len(color_list):
        return ws
    ws.cell(row=m+2, column=2*b+1).value = f"dummy car"
    ws.cell(row=m+2, column=2*b+1).fill = color_list[m]
    ws.cell(row=m+2, column=2*b+1).alignment = center_alignment
    ws.cell(row=m+2, column=2*b+2).value = f"{LP_list[m]} → {DP_list[m]}"
    ws.cell(row=m+2, column=2*b+2).alignment = center_alignment
        
    return ws


def fit_cell_size(ws, a, b):
    for i in range(2*a+1):
        if i % 2 == 1:
            ws.row_dimensions[i+1].height = 20
        else:
            ws.row_dimensions[i+1].height = 50
    for i in range(2*b+1):
        if i % 2 == 1:
            ws.column_dimensions[get_column_letter(i+1)].width = 5
        else:
            ws.column_dimensions[get_column_letter(i+1)].width = 10
    
    return


def output_solution_info(ws, penalty_sol):
    weight_list = [MASTER.w1, MASTER.w2, MASTER.w3, MASTER.w4, MASTER.w5, MASTER.w6, MASTER.w7]
    
    for i in range(8):
        if i == 0:
            ws.cell(row=i+1, column=2).value = "weight"
            ws.cell(row=i+1, column=3).value = "solution value"
            continue
        ws.cell(row=i+1, column=1).value = f"penalty {i}"
        ws.cell(row=i+1, column=2).value = weight_list[i-1]
        ws.cell(row=i+1, column=3).value = penalty_sol[i-1]
        
    return


def visualize_solution(sol, sola, solb, penalty_sol, a, b, m, LP_list, DP_list, Amount_list, M_h_bar):
    wb = excel.Workbook()
    
    # 積み込み時の情報
    ws = wb.active
    ws.title = "積み込み"
    ws = paint_cell(a, b, sol, ws, m)
    print_edge_in(a, b, sola, ws)
    add_annotation(ws, a, b, m, LP_list, DP_list, Amount_list, M_h_bar)
    point_enter_cell(a,b,"in",ws)
    fit_cell_size(ws, a, b)
    
    # 搬出時の情報
    wb.create_sheet(title="搬出")
    ws = wb["搬出"]
    ws = paint_cell(a, b, sol, ws, m)
    print_edge_out(a, b, solb, ws)
    add_annotation(ws, a, b, m, LP_list, DP_list, Amount_list, M_h_bar)
    point_enter_cell(a,b,"out",ws)
    fit_cell_size(ws, a, b)
    
    # 定性情報
    wb.create_sheet(title="問題情報")
    ws = wb["問題情報"]
    output_solution_info(ws, penalty_sol)
    
    # 結果の保存
    now_time = datetime.datetime.now()
    now_time_str = now_time.strftime("%Y%m%d_%H%M%S")
    wb.save(f"results/{now_time_str}_{MASTER.input_property}_{MASTER.input_DK}dk.xlsx")
    
    return


if __name__ == "__main__":
    a = 5 # input row of brock
    b = 5 # input column of brock
    # main(a,b)

