import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import make_instance_tool
import MASTER

# color setting
yellow_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFFF00')
red_fill = excel.styles.PatternFill(patternType='solid', fgColor='FF0000')
blue_fill = excel.styles.PatternFill(patternType='solid', fgColor='0000FF')
green_fill = excel.styles.PatternFill(patternType='solid', fgColor='00FF00')
pink_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFC0CB')
skyblue_fill = excel.styles.PatternFill(patternType='solid', fgColor='87CEEB')
rime_fill = excel.styles.PatternFill(patternType='solid', fgColor='FAEBD7')
gray_fill = excel.styles.PatternFill(patternType='solid', fgColor='808080')
purple_fill = excel.styles.PatternFill(patternType='solid', fgColor='800080')
orange_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFA500')
yellow_green_fill = excel.styles.PatternFill(patternType='solid', fgColor='9ACD32')
pink_gold_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFD700')
wine_red_fill = excel.styles.PatternFill(patternType='solid', fgColor='800000')
marine_blue_fill = excel.styles.PatternFill(patternType='solid', fgColor='000080')
passion_pink_fill = excel.styles.PatternFill(patternType='solid', fgColor='FF1493')

color_list = [yellow_fill, red_fill, blue_fill, green_fill, pink_fill, skyblue_fill, rime_fill, gray_fill, purple_fill, orange_fill, yellow_green_fill, pink_gold_fill, wine_red_fill, marine_blue_fill, passion_pink_fill]

block_border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))

ramp_border = Border(top=Side(style='mediumDashed', color='FF0000'), 
                bottom=Side(style='mediumDashed', color='FF0000'), 
                left=Side(style='mediumDashed', color='FF0000'),
                right=Side(style='mediumDashed', color='FF0000'))

center_alignment = Alignment(horizontal="centerContinuous", vertical="center" , wrap_text=True)
red_font = excel.styles.fonts.Font(color='FF0000')
brue_font = excel.styles.fonts.Font(color='0000FF')


def paint_cell(a, b, sol, ws, m):
    enter_block, exit_block = make_instance_tool.get_ramp_block(a, b)
    
    for i in range(a):
        for j in range(b):
            ws.cell(row=2*i+1, column=2*j+1).value = i*b+j
            ws.cell(row=2*i+1, column=2*j+1).border = block_border
            ws.cell(row=2*i+1, column=2*j+1).alignment = center_alignment
            
            if i*b+j == enter_block or i*b+j == exit_block:
                continue
            else:
                ws.cell(row=2*i+1, column=2*j+1).fill = color_list[sol[i*b+j]]
    
    # dummy car
    ws.cell(row=2*(exit_block//b)+1, column=2*(exit_block%b)+1).fill = color_list[m]
    
    return ws


def point_enter_cell(a, b, in_out_flag, ws):
    enter_block, exit_block = make_instance_tool.get_ramp_block(a, b)
    ws.cell(row=2*(enter_block//b)+1, column=2*(enter_block%b)+1).border = ramp_border
    
    return ws


def print_edge_in(a, b, sola, ws):
    for edge in sola:
        if edge[0] + 1 == edge[1]:
            # print(f"edge →: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = (cell_contents or "") + "→"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge ←: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = (cell_contents or "") + "←"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    return


def print_edge_out(a, b, solb, ws):
    for edge in solb:
        if edge[0] + 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = (cell_contents or "") + "→"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value
            if cell_contents is not None:
                cell_contents = cell_contents + "\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = (cell_contents or "") + "←"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    
    return


def add_annotation(ws, a, b, m, LP_list, DP_list):
    ws.cell(row=1, column=2*b+1).value = f"car"
    ws.cell(row=1, column=2*b+1).alignment = center_alignment
    ws.cell(row=1, column=2*b+2).value = f"LP → DP"
    ws.cell(row=1, column=2*b+2).alignment = center_alignment
    
    for i in range(m):
        ws.cell(row=i+2, column=2*b+1).value = f"car {i}"
        ws.cell(row=i+2, column=2*b+1).fill = color_list[i]
        ws.cell(row=i+2, column=2*b+1).alignment = center_alignment
        ws.cell(row=i+2, column=2*b+2).value = f"{LP_list[i]} → {DP_list[i]}"
        ws.cell(row=i+2, column=2*b+2).alignment = center_alignment
    
    # dummy car
    ws.cell(row=m+2, column=2*b+1).value = f"dummy car"
    ws.cell(row=m+2, column=2*b+1).fill = color_list[m]
    ws.cell(row=m+2, column=2*b+1).alignment = center_alignment
    ws.cell(row=m+2, column=2*b+2).value = f"{LP_list[m]} → {DP_list[m]}"
    ws.cell(row=m+2, column=2*b+2).alignment = center_alignment
        
    return ws


def fit_cell_size(ws, a, b):
    for i in range(2*a+1):
        if i % 2 == 1:
            ws.row_dimensions[i+1].height = 20
        else:
            ws.row_dimensions[i+1].height = 50
    for i in range(2*b+1):
        if i % 2 == 1:
            ws.column_dimensions[get_column_letter(i+1)].width = 5
        else:
            ws.column_dimensions[get_column_letter(i+1)].width = 10
    
    return


def visualize_solution(sol, sola, solb, a, b, m, LP_list, DP_list, model_name: str):
    wb = excel.Workbook()
    
    # 積み込み時の情報
    ws = wb.active
    ws.title = "積み込み"
    ws = paint_cell(a, b, sol, ws, m)
    print_edge_in(a, b, sola, ws)
    add_annotation(ws, a, b, m, LP_list, DP_list)
    point_enter_cell(a,b,"in",ws)
    fit_cell_size(ws, a, b)
    
    # 搬出時の情報
    wb.create_sheet(title="搬出")
    ws = wb["搬出"]
    ws = paint_cell(a, b, sol, ws, m)
    print_edge_out(a, b, solb, ws)
    add_annotation(ws, a, b, m, LP_list, DP_list)
    point_enter_cell(a,b,"out",ws)
    fit_cell_size(ws, a, b)
    
    
    model_name = "results"
    if MASTER.potential_flag == 0:
        model_name += "_Normal"
    else:
        model_name += "_Potential"
    if MASTER.next_block_flag == 1:
        model_name += "_NextBlock"

    wb.save(f"{model_name}.xlsx")
    
    return


if __name__ == "__main__":
    a = 5 # input row of brock
    b = 5 # input column of brock
    # main(a,b)

