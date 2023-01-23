import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import test, flow

# color setting
red_fill = excel.styles.PatternFill(patternType='solid', fgColor='FF0000')
blue_fill = excel.styles.PatternFill(patternType='solid', fgColor='0000FF')
green_fill = excel.styles.PatternFill(patternType='solid', fgColor='00FF00')
yellow_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFFF00')
pink_fill = excel.styles.PatternFill(patternType='solid', fgColor='FFC0CB')
skyblue_fill = excel.styles.PatternFill(patternType='solid', fgColor='87CEEB')

color_list = [red_fill, blue_fill, green_fill, yellow_fill, pink_fill, skyblue_fill]

block_border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))

center_alignment = Alignment(horizontal="centerContinuous", vertical="center" , wrap_text=True)
red_font = excel.styles.fonts.Font(color='FF0000')
brue_font = excel.styles.fonts.Font(color='0000FF')


def paint_cell(a, b, sol, ws):
    for i in range(a):
        for j in range(b):
            # print(f"{i*a+j}: {sol[i*a+j]}")
            ws.cell(row=2*i+1, column=2*j+1).value = i*b+j
            ws.cell(row=2*i+1, column=2*j+1).border = block_border
            ws.cell(row=2*i+1, column=2*j+1).alignment = center_alignment
            
            if sol[i*b+j] == 0:
                pass
            elif sol[i*b+j] == 1:
                ws.cell(row=2*i+1, column=2*j+1).fill = red_fill
            elif sol[i*b+j] == 2:
                ws.cell(row=2*i+1, column=2*j+1).fill = blue_fill
            elif sol[i*b+j] == 3:
                ws.cell(row=2*i+1, column=2*j+1).fill = green_fill
            elif sol[i*b+j] == 4:
                ws.cell(row=2*i+1, column=2*j+1).fill = yellow_fill
            else:
                pass
            
    return ws


def print_edge_in(a, b, sola, ws):
    for edge in sola:
        if edge[0] + 1 == edge[1]:
            # print(f"edge: {edge}")
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = "→\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge: {edge}")
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = "←\n"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    return


def print_edge_out(a, b, solb, ws):
    for edge in solb:
        if edge[0] + 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).value = (cell_contents or "") + "→"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)+2).alignment = center_alignment
        elif edge[0] - 1 == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).value = (cell_contents or "") + "←"
            ws.cell(row=2*(edge[0]//b)+1, column=2*(edge[0]%b)).alignment = center_alignment
        elif edge[0] + b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↓"
            ws.cell(row=2*(edge[0]//b)+2, column=2*(edge[0]%b)+1).alignment = center_alignment
        elif edge[0] - b == edge[1]:
            # print(f"edge: {edge}")
            cell_contents = ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).value = (cell_contents or "") + "↑"
            ws.cell(row=2*(edge[0]//b), column=2*(edge[0]%b)+1).alignment = center_alignment
    
    return


def get_LP_DP_list():
    with open("sample_data.dat", 'r') as f:
    # with open("small.dat", 'r') as f:
        lines = [line.rstrip() for line in f.readlines()]
    lines = [line for line in lines if line[0] != '#']
    items = lines.pop(0).split(' ')
    m, n = int(items[0]),  int(items[1])
    M = {i for i in range(1,m+1)}
    M_p = {i for i in range(1,m+2)}
    V = {i for i in range(1,n+1)}
    V_p = {i for i in range(0,n+2)}
    items = lines.pop(0).split(' ')
    p = {i+1: int(v) for i, v in enumerate(items)}
    items = lines.pop(0).split(' ')
    o = {i+1: int(v) for i, v in enumerate(items)}
    items = lines.pop(0).split(' ')
    d = {i+1: int(v) for i, v in enumerate(items)}
    items = lines.pop(0).split(' ')
    
    # print(f"o: {o}")
    # print(f"d: {d}")
        
    return o, d


def add_annotation(ws, a, b, m):
    LP_list, DP_list = get_LP_DP_list()
    
    ws.cell(row=1, column=2*b+1).value = f"car"
    ws.cell(row=1, column=2*b+1).alignment = center_alignment
    ws.cell(row=1, column=2*b+2).value = f"LP → DP"
    ws.cell(row=1, column=2*b+2).alignment = center_alignment
    
    for i in range(m):
        ws.cell(row=i+2, column=2*b+1).value = f"car {i+1}"
        ws.cell(row=i+2, column=2*b+1).fill = color_list[i]
        ws.cell(row=i+2, column=2*b+1).alignment = center_alignment
        ws.cell(row=i+2, column=2*b+2).value = f"{LP_list[i+1]} → {DP_list[i+1]}"
        ws.cell(row=i+2, column=2*b+2).alignment = center_alignment
        
    return ws


def fit_cell_size(ws, a, b):
    for i in range(2*a+1):
        ws.row_dimensions[i+1].height = 32
    for i in range(2*b+1):
        ws.column_dimensions[get_column_letter(i+1)].width = 8
    

# def for_test(a,b):
#     sol, sola, solb = test.main()
#     sol[0] = 0 # 0 is not used
    
#     wb = excel.Workbook()
#     wb.save("results_tree.xlsx")
#     ws = wb.active
    
#     ws = paint_cell(a, b, sol, ws)
#     print_edge_in(a, b, sola, ws)
#     print_edge_out(a, b, solb, ws)
    
#     add_annotation(ws, a, b, 4)
#     fit_cell_size(ws, a, b)
    
#     wb.save("results_tree.xlsx")
    
#     return


# def for_flow(a,b):
#     sol, sola, solb = flow.main()
#     sol[0] = 0 # 0 is not used
    
#     wb = excel.Workbook()
#     wb.save("results_flow.xlsx")
#     ws = wb.active
    
#     ws = paint_cell(a, b, sol, ws)
#     print_edge_in(a, b, sola, ws)
#     print_edge_out(a, b, solb, ws)
    
#     add_annotation(ws, a, b, 4)
#     fit_cell_size(ws, a, b)
    
#     wb.save("results_flow.xlsx")
    
#     # print(len(sola), len(solb))
#     # print(sola)
#     # print(solb)
    
#     return

    
# def main(a,b):
#     for_flow(a,b)
#     for_test(a,b)
    
#     return


def visualize_solution(sol, sola, solb, a, b, model_name: str):
    sol[0] = 0 # 0 is not used
    
    wb = excel.Workbook()
    wb.save("results_flow.xlsx")
    ws = wb.active
    
    ws = paint_cell(a, b, sol, ws)
    print_edge_in(a, b, sola, ws)
    print_edge_out(a, b, solb, ws)
    
    add_annotation(ws, a, b, 4)
    fit_cell_size(ws, a, b)
    
    wb.save(f"results_{model_name}.xlsx")
    
    return


if __name__ == "__main__":
    a = 5 # input row of brock
    b = 5 # input column of brock
    # main(a,b)

